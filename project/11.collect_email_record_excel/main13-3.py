# 사이트에서 이메일 수집

from openpyxl import Workbook, load_workbook
import requests
import re

url1 = 'https://news.v.daum.net/v/20220627172955380'
url2 = 'https://n.news.naver.com/article/448/0000364455?cds=news_media_pc'

headers = {
    'User-Agent' : 'Mozilla/5.0',
    'Content-Type' : 'text/html; charset=utr-8'
}

response = requests.get(url1, headers = headers)
results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)
results = list(set(results))

print(results)

try:
    wb = load_workbook(r'project\13.collect_email_record_excel\cere.xlsx', data_only=True)
    sheet = wb.active
    
except:
    wb = Workbook()
    sheet = wb.active
    
for result in results:
    sheet.append([result])
    
wb.save(r'project\13.collect_email_record_excel\cere.xlsx')